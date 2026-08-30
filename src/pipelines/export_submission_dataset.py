import json
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# ============================================================
# PATHS
# ============================================================

BASE = Path("data")

PROCESSED = BASE / "processed"

ENTITIES_FILE = (
    PROCESSED / "ai_entities_classified.json"
)

PAPERS_FILE = (
    PROCESSED / "research_papers_extracted.json"
)

RESOLUTION_FILE = (
    PROCESSED / "entity_resolution.json"
)

FRESH_FILE = (
    PROCESSED / "fresh_ai_data.json"
)

OUTPUT_FILE = (
    PROCESSED / "graphone_submission_dataset.xlsx"
)


# ============================================================
# HELPERS
# ============================================================

def load_json(path):
    if not path.exists():
        raise FileNotFoundError(
            f"Required file not found: {path}"
        )

    with open(
        path,
        "r",
        encoding="utf-8"
    ) as f:
        return json.load(f)


def ensure_list(data):
    """
    Converts common JSON wrapper structures into a list.
    """

    if isinstance(data, list):
        return data

    if isinstance(data, dict):

        for key in (
            "items",
            "records",
            "entities",
            "papers",
            "data"
        ):

            value = data.get(key)

            if isinstance(value, list):
                return value

    return []


def safe(value):
    """
    Convert nested structures into readable strings
    for spreadsheet cells.
    """

    if value is None:
        return ""

    if isinstance(value, list):
        return ", ".join(
            str(x)
            for x in value
        )

    if isinstance(value, dict):
        return json.dumps(
            value,
            ensure_ascii=False
        )

    return value


def write_sheet(
    workbook,
    title,
    headers,
    rows
):

    ws = workbook.create_sheet(title)

    ws.append(headers)

    for row in rows:
        ws.append(
            [
                safe(value)
                for value in row
            ]
        )

    # Header formatting
    for cell in ws[1]:
        cell.font = cell.font.copy(
            bold=True
        )

    ws.freeze_panes = "A2"

    ws.auto_filter.ref = ws.dimensions

    # Automatic column width
    for column_cells in ws.columns:

        max_length = 0

        column_letter = get_column_letter(
            column_cells[0].column
        )

        for cell in column_cells:

            value = str(
                cell.value
                if cell.value is not None
                else ""
            )

            max_length = max(
                max_length,
                len(value)
            )

        ws.column_dimensions[
            column_letter
        ].width = min(
            max(max_length + 2, 12),
            60
        )

    return ws


# ============================================================
# MAIN
# ============================================================

def main():

    print("=" * 70)
    print("GRAPHONE SUBMISSION DATASET EXPORTER")
    print("=" * 70)

    # --------------------------------------------------------
    # LOAD DATA
    # --------------------------------------------------------

    print()
    print("Loading processed datasets...")

    entities_raw = load_json(
        ENTITIES_FILE
    )

    papers_raw = load_json(
        PAPERS_FILE
    )

    resolution_raw = load_json(
        RESOLUTION_FILE
    )

    fresh_raw = load_json(
        FRESH_FILE
    )

    entities = ensure_list(
        entities_raw
    )

    papers = ensure_list(
        papers_raw
    )

    # Resolution can be either a list or a dict
    if isinstance(resolution_raw, dict):

        resolution = resolution_raw

        # Actual structure of entity_resolution.json
        links = resolution.get(
            "paperEntityLinks",
            []
        )

        if not isinstance(links, list):
            links = []

    else:

        resolution = {}

        links = ensure_list(
            resolution_raw
        )

    # --------------------------------------------------------
    # FRESH DATA
    # --------------------------------------------------------

    if not isinstance(
        fresh_raw,
        dict
    ):
        fresh_raw = {}

    news_section = fresh_raw.get(
        "news",
        {}
    )

    jobs_section = fresh_raw.get(
        "jobs",
        {}
    )

    if not isinstance(
        news_section,
        dict
    ):
        news_section = {}

    if not isinstance(
        jobs_section,
        dict
    ):
        jobs_section = {}

    news = news_section.get(
        "items",
        []
    )

    jobs = jobs_section.get(
        "items",
        []
    )

    if not isinstance(news, list):
        news = []

    if not isinstance(jobs, list):
        jobs = []

    # --------------------------------------------------------
    # CLASSIFY ENTITIES
    # --------------------------------------------------------

    startups = []
    products = []

    for entity in entities:

        entity_type = str(
            entity.get(
                "entityType",
                ""
            )
        ).upper()

        if entity_type == "STARTUP":

            startups.append(entity)

        elif entity_type == "PRODUCT":

            products.append(entity)

    # --------------------------------------------------------
    # CREATE WORKBOOK
    # --------------------------------------------------------

    workbook = Workbook()

    # Remove default sheet
    default_sheet = workbook.active

    workbook.remove(
        default_sheet
    )

    # ========================================================
    # TAB 1 — STARTUPS
    # ========================================================

    startup_headers = [
        "Entity Type",
        "Entity Name",
        "Full Name",
        "Description",
        "GitHub URL",
        "Homepage",
        "Owner",
        "Owner Type",
        "Language",
        "Stars",
        "Forks",
        "Open Issues",
        "License",
        "Created At",
        "Updated At",
        "Topics",
        "Collected At",
        "AI Relevant",
        "Confidence",
        "Reason",
        "Classification Method",
        "Source Verified"
    ]

    startup_rows = []

    for entity in startups:

        startup_rows.append([
            entity.get("entityType"),
            entity.get("entityName"),
            entity.get("fullName"),
            entity.get("description"),
            entity.get("githubUrl"),
            entity.get("homepage"),
            entity.get("owner"),
            entity.get("ownerType"),
            entity.get("language"),
            entity.get("stars"),
            entity.get("forks"),
            entity.get("openIssues"),
            entity.get("license"),
            entity.get("createdAt"),
            entity.get("updatedAt"),
            entity.get("topics"),
            entity.get("collectedAt"),
            entity.get("isAIRelevant"),
            entity.get("confidence"),
            entity.get("reason"),
            entity.get("classificationMethod"),
            entity.get("sourceVerified")
        ])

    write_sheet(
        workbook,
        "Startups",
        startup_headers,
        startup_rows
    )

    # ========================================================
    # TAB 2 — PRODUCTS
    # ========================================================

    product_headers = startup_headers

    product_rows = []

    for entity in products:

        product_rows.append([
            entity.get("entityType"),
            entity.get("entityName"),
            entity.get("fullName"),
            entity.get("description"),
            entity.get("githubUrl"),
            entity.get("homepage"),
            entity.get("owner"),
            entity.get("ownerType"),
            entity.get("language"),
            entity.get("stars"),
            entity.get("forks"),
            entity.get("openIssues"),
            entity.get("license"),
            entity.get("createdAt"),
            entity.get("updatedAt"),
            entity.get("topics"),
            entity.get("collectedAt"),
            entity.get("isAIRelevant"),
            entity.get("confidence"),
            entity.get("reason"),
            entity.get("classificationMethod"),
            entity.get("sourceVerified")
        ])

    write_sheet(
        workbook,
        "Products",
        product_headers,
        product_rows
    )

    # ========================================================
    # TAB 3 — RESEARCH PAPERS
    # ========================================================

    paper_headers = [
        "Record Type",
        "Title",
        "Authors",
        "Abstract",
        "Source URL",
        "GitHub URL",
        "GitHub Stars"
    ]

    paper_rows = []

    for paper in papers:

        paper_rows.append([
            paper.get("recordType"),
            paper.get("title"),
            paper.get("authors"),
            paper.get("abstract"),
            paper.get("sourceUrl"),
            paper.get("githubUrl"),
            paper.get("githubStars")
        ])

    write_sheet(
        workbook,
        "Research Papers",
        paper_headers,
        paper_rows
    )

    # ========================================================
    # TAB 4 — JOBS
    # ========================================================

    job_headers = [
        "Source",
        "URL",
        "HTTP Status",
        "Accessible",
        "Checked At"
    ]

    job_rows = []

    for job in jobs:

        job_rows.append([
            job.get("source"),
            job.get("url"),
            job.get("status"),
            job.get("accessible"),
            job.get("checkedAt")
        ])

    write_sheet(
        workbook,
        "Jobs",
        job_headers,
        job_rows
    )

    # ========================================================
    # TAB 5 — NEWS
    # ========================================================

    news_headers = [
        "Type",
        "Source",
        "Title",
        "URL",
        "Published At",
        "Summary",
        "Collected At"
    ]

    news_rows = []

    for article in news:

        news_rows.append([
            article.get("type"),
            article.get("source"),
            article.get("title"),
            article.get("url"),
            article.get("publishedAt"),
            article.get("summary"),
            article.get("collectedAt")
        ])

    write_sheet(
        workbook,
        "News",
        news_headers,
        news_rows
    )

    # ========================================================
    # TAB 6 — ENTITY MAPPING LOG
    # ========================================================

    mapping_headers = [
        "Paper Source URL",
        "Paper Title",
        "Entity Name",
        "Entity GitHub URL",
        "Match Type",
        "Match Score",
        "Resolution Method",
        "Status"
    ]

    mapping_rows = []

    for link in links:

        mapping_rows.append([
            link.get(
                "paperSourceUrl",
                link.get(
                    "sourceUrl"
                )
            ),

            link.get(
                "paperTitle",
                link.get(
                    "title"
                )
            ),

            link.get(
                "entityName"
            ),

            link.get(
                "githubUrl"
            ),

            link.get(
                "matchType"
            ),

            link.get(
                "score",
                link.get(
                    "matchScore"
                )
            ),

            link.get(
                "method",
                link.get(
                    "resolutionMethod"
                )
            ),

            "RESOLVED"
        ])

    # Add unresolved papers if resolution contains them
    unresolved = resolution.get(
        "unresolvedPapers",
        []
    )

    if isinstance(
        unresolved,
        list
    ):

        for item in unresolved:

            if isinstance(
                item,
                dict
            ):

                mapping_rows.append([
                    item.get(
                        "sourceUrl"
                    ),
                    item.get(
                        "title"
                    ),
                    "",
                    "",
                    "",
                    "",
                    "",
                    "UNRESOLVED"
                ])

    write_sheet(
        workbook,
        "Entity Mapping Log",
        mapping_headers,
        mapping_rows
    )

    # ========================================================
    # SAVE
    # ========================================================

    OUTPUT_FILE.parent.mkdir(
        parents=True,
        exist_ok=True
    )

    workbook.save(
        OUTPUT_FILE
    )

    # ========================================================
    # SUMMARY
    # ========================================================

    print()
    print("=" * 70)
    print("EXPORT COMPLETE")
    print("=" * 70)

    print(
        f"Startups            : {len(startups)}"
    )

    print(
        f"Products            : {len(products)}"
    )

    print(
        f"Research papers     : {len(papers)}"
    )

    print(
        f"Job source records  : {len(jobs)}"
    )

    print(
        f"Fresh news          : {len(news)}"
    )

    print(
        f"Mapping records     : {len(mapping_rows)}"
    )

    print()
    print(
        f"OUTPUT: {OUTPUT_FILE}"
    )

    print()
    print(
        "Workbook contains 6 tabs:"
    )

    print(
        "1. Startups"
    )

    print(
        "2. Products"
    )

    print(
        "3. Research Papers"
    )

    print(
        "4. Jobs"
    )

    print(
        "5. News"
    )

    print(
        "6. Entity Mapping Log"
    )

    print("=" * 70)


if __name__ == "__main__":
    main()